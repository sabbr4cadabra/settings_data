# percorrer toda a base de dados
# para cada item
# para cada item, ver se o bairro ja existe em uma aba, se não existir, criar aquela aba
# copiar os valores daquela linha e colocar na aba do bairro correspondente

from openpyxl import load_workbook
from copy import copy

def criar_aba(bairro, arquivo_bairros):
    if bairro not in arquivo_bairros.sheetnames:
        arquivo_bairros.create_sheet(bairro)
        nova_aba = arquivo_bairros[bairro]
        nova_aba['A1'].value = 'Data de Nascimento'
        nova_aba['B1'].value = 'Pessoa'
        nova_aba['C1'].value = 'Bairro'

def transferir_informacoes_aba(aba_origem, aba_destino, linha_origem):
    linha_destino = aba_destino.max_row + 1
    for coluna in range(1, 4):
        celula_origem = aba_origem.cell(row=linha_origem, column=coluna)
        celula_destino = aba_destino.cell(row=linha_destino, column=coluna)
        celula_destino.value = celula_origem.value
        celula_destino._style = copy(celula_origem._style)
        
arquivo_bairros = load_workbook('Bairros.xlsx')

print(arquivo_bairros.sheetnames)

aba_basedados = arquivo_bairros['Base de Dados']

ultima_linha = aba_basedados.max_row
print(ultima_linha)

for linha in range(2, ultima_linha + 1):
    bairro = aba_basedados[f'C{linha}'].value
    if not bairro:
        break
    
    # Criar uma aba para o bairro, se não existir
    criar_aba(bairro, arquivo_bairros)
    
    # Transferir as informações para a aba correspondente
    aba_origem = aba_basedados  # Definir a aba de origem como 'Base de Dados'
    aba_destino = arquivo_bairros[bairro]  # Definir a aba de destino
    transferir_informacoes_aba(aba_origem, aba_destino, linha)  # Passar a linha atual
    
arquivo_bairros.save('Bairros2.xlsx')
